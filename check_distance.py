import openpyxl
import numpy as np

def calculate_distance(x,y):
    return np.sqrt(((x[0]-y[0])*110.574)**2 + ((x[1]-y[1])*111.32)**2)

#wb = openpyxl.load_workbook('Address_Without_MRT.xlsx')
#wb2 = openpyxl.load_workbook('MRT_List.xlsx')
wb = openpyxl.load_workbook('Street_Name_List.xlsx')
wb2 = openpyxl.load_workbook('Shopping_Mall_List.xlsx')
sheet = wb['Sheet1']
sheet2 = wb2['Sheet1']
for row in range (2, sheet.max_row +1):
	Distance = 200
	MRT = 0
	if sheet['B'+str(row)].value is not None:
		for k in range (2, sheet2.max_row + 1):
			x1 = sheet['B'+str(row)].value
			x2 = float(sheet2['B'+str(k)].value)
			y1 = sheet['C'+str(row)].value
			y2 = float(sheet2['C'+str(k)].value)
			#print (type(x1),type(x2),y1,y2)
			Dis_Temp = calculate_distance([x1,y1],[x2,y2])
			if Dis_Temp < Distance:
				Distance = Dis_Temp
				MRT = k
			else:
				continue
	else:
		continue
	print(MRT)
	print(sheet2['A'+str(MRT)].value)
	sheet['F'+str(row)].value = sheet2['A'+str(MRT)].value

#wb.save('Address_Without_MRT.xlsx')
wb.save('Street_Name_List.xlsx')
print('Done.')